#!/usr/bin/env python3
"""
🐣 Система уведомлений о доставке — Азовский Инкубатор
=====================================================

Читает Excel-таблицу с клиентами и рассылает SMS-уведомления
о доставке цыплят в указанное время и место.

Запуск:
  DRY RUN (тест):  python3 notifier.py --file clients.xlsx --dry-run
  БОЕВОЙ:          python3 notifier.py --file clients.xlsx --provider smsru --api-key XXX

API-сервер (для веб-панели):
  python3 notifier.py --server --port 8080

v1.0 — 15.04.2026
"""
import os
import sys
import json
import re
import argparse
import time
from datetime import datetime
from http.server import HTTPServer, SimpleHTTPRequestHandler
from urllib.parse import parse_qs, urlparse
import io

# Добавляем путь к модулям
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from sms_providers import get_provider, DryRunProvider

# ============================================================
# 📊 Чтение Excel
# ============================================================

def normalize_phone(raw: str) -> str:
    """Нормализует номер телефона в формат 79XXXXXXXXX."""
    if not raw:
        return ""
    raw = str(raw).strip()
    # Убираем всё кроме цифр и +
    digits = re.sub(r'[^\d+]', '', raw)
    # Убираем +
    digits = digits.replace('+', '')
    
    if digits.startswith('8') and len(digits) == 11:
        digits = '7' + digits[1:]
    elif digits.startswith('9') and len(digits) == 10:
        digits = '7' + digits
    elif not digits.startswith('7'):
        return ""  # Не российский номер
    
    if len(digits) != 11:
        return ""
    
    return digits


def read_excel(filepath: str) -> list:
    """Читает Excel и возвращает список клиентов.
    
    Ожидаемые колонки (в любом порядке, ищем по ключевым словам в заголовке):
    - Имя / Клиент / ФИО / Name
    - Телефон / Phone / Номер
    - Заказ / Товар / Order / Породы
    - Адрес / Город / Address / Место
    - Время / Дата / Time / Date / Доставка
    """
    import openpyxl
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    
    # Ищем строку с заголовками (первая строка с хотя бы 2 заполненными ячейками)
    header_row = None
    headers = {}
    
    column_patterns = {
        'name': ['имя', 'клиент', 'фио', 'name', 'заказчик', 'покупатель'],
        'phone': ['телефон', 'phone', 'номер', 'тел', 'моб'],
        'order': ['заказ', 'товар', 'order', 'пород', 'птица', 'кол-во', 'количество'],
        'address': ['адрес', 'город', 'address', 'место', 'точка', 'населённый', 'населенный'],
        'time': ['время', 'дата', 'time', 'date', 'доставка', 'когда', 'прибытие'],
    }
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=False), start=1):
        filled = sum(1 for c in row if c.value)
        if filled >= 2:
            for col_idx, cell in enumerate(row):
                if cell.value:
                    val = str(cell.value).lower().strip()
                    for field, patterns in column_patterns.items():
                        if any(p in val for p in patterns):
                            headers[field] = col_idx
            
            if 'phone' in headers:  # Минимум — нужен телефон
                header_row = row_idx
                break
    
    if not header_row:
        raise ValueError("Не найдена строка с заголовками. Нужна колонка 'Телефон'.")
    
    # Читаем данные
    clients = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=False):
        phone_raw = str(row[headers.get('phone', 0)].value or "")
        phone = normalize_phone(phone_raw)
        
        if not phone:
            continue
        
        client = {
            'name': str(row[headers.get('name', 0)].value or "Клиент") if 'name' in headers else "Клиент",
            'phone': phone,
            'phone_raw': phone_raw,
            'order': str(row[headers.get('order', 0)].value or "") if 'order' in headers else "",
            'address': str(row[headers.get('address', 0)].value or "") if 'address' in headers else "",
            'time': str(row[headers.get('time', 0)].value or "") if 'time' in headers else "",
        }
        clients.append(client)
    
    return clients


# ============================================================
# 📝 Шаблон сообщения
# ============================================================

DEFAULT_TEMPLATE = (
    "Здравствуйте, {name}! 🐣\n"
    "Ваш заказ ({order}) будет доставлен:\n"
    "📍 {address}\n"
    "🕐 {time}\n"
    "Азовский Инкубатор | 8-978-XXX-XX-XX"
)

def render_message(template: str, client: dict) -> str:
    """Подставляет данные клиента в шаблон."""
    msg = template
    for key, val in client.items():
        placeholder = "{" + key + "}"
        msg = msg.replace(placeholder, val if val else "—")
    return msg.strip()


# ============================================================
# 📤 Рассылка
# ============================================================

def send_notifications(clients: list, template: str, provider, delay: float = 1.0) -> list:
    """Рассылает SMS всем клиентам.
    
    Args:
        clients: список клиентов из read_excel()
        template: шаблон сообщения
        provider: SMS-провайдер
        delay: задержка между SMS (секунды)
    
    Returns:
        список результатов [{client, message, result}]
    """
    results = []
    total = len(clients)
    
    for i, client in enumerate(clients, 1):
        message = render_message(template, client)
        
        print(f"[{i}/{total}] 📱 {client['name']} ({client['phone']})...")
        
        result = provider.send(client['phone'], message)
        
        results.append({
            "index": i,
            "client": client,
            "message": message,
            "result": result,
            "timestamp": datetime.now().isoformat()
        })
        
        status = "✅" if result['success'] else "❌"
        print(f"  {status} {result.get('error', 'OK')}")
        
        if i < total:
            time.sleep(delay)
    
    # Итоги
    ok = sum(1 for r in results if r['result']['success'])
    fail = total - ok
    print(f"\n📊 Итого: {ok} отправлено, {fail} ошибок из {total}")
    
    return results


# ============================================================
# 🌐 HTTP-сервер (для веб-панели)
# ============================================================

_server_state = {
    "clients": [],
    "results": [],
    "provider_name": "dryrun",
    "provider_config": {},
    "template": DEFAULT_TEMPLATE,
}


class NotifierHandler(SimpleHTTPRequestHandler):
    """API для веб-панели."""
    
    def do_GET(self):
        parsed = urlparse(self.path)
        
        if parsed.path == "/api/status":
            self._json_response({
                "clients_count": len(_server_state["clients"]),
                "results_count": len(_server_state["results"]),
                "provider": _server_state["provider_name"],
                "template": _server_state["template"],
            })
        elif parsed.path == "/api/clients":
            self._json_response(_server_state["clients"])
        elif parsed.path == "/api/results":
            self._json_response(_server_state["results"])
        elif parsed.path == "/api/template":
            self._json_response({"template": _server_state["template"]})
        elif parsed.path == "/" or parsed.path == "/index.html":
            # Отдаём dashboard.html
            dashboard_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "dashboard.html")
            if os.path.exists(dashboard_path):
                with open(dashboard_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                self.send_response(200)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.end_headers()
                self.wfile.write(content.encode('utf-8'))
            else:
                self._json_response({"error": "dashboard.html not found"}, 404)
        else:
            self._json_response({"error": "Not found"}, 404)
    
    def do_POST(self):
        parsed = urlparse(self.path)
        content_length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(content_length)
        
        if parsed.path == "/api/upload":
            self._handle_upload(body)
        elif parsed.path == "/api/template":
            self._handle_template(body)
        elif parsed.path == "/api/send":
            self._handle_send(body)
        elif parsed.path == "/api/preview":
            self._handle_preview(body)
        elif parsed.path == "/api/provider":
            self._handle_provider(body)
        else:
            self._json_response({"error": "Not found"}, 404)
    
    def _handle_upload(self, body):
        """Загрузка Excel-файла."""
        try:
            # Ищем boundary для multipart
            content_type = self.headers.get('Content-Type', '')
            
            if 'multipart/form-data' in content_type:
                # Парсим multipart вручную (без cgi)
                boundary = content_type.split('boundary=')[1].strip()
                parts = body.split(f'--{boundary}'.encode())
                
                for part in parts:
                    if b'filename=' in part:
                        # Извлекаем файл
                        header_end = part.find(b'\r\n\r\n')
                        file_data = part[header_end + 4:]
                        # Убираем trailing \r\n
                        if file_data.endswith(b'\r\n'):
                            file_data = file_data[:-2]
                        
                        # Сохраняем во временный файл
                        tmp_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "_upload.xlsx")
                        with open(tmp_path, 'wb') as f:
                            f.write(file_data)
                        
                        # Читаем
                        clients = read_excel(tmp_path)
                        _server_state["clients"] = clients
                        
                        # Чистим
                        os.remove(tmp_path)
                        
                        self._json_response({
                            "success": True,
                            "clients_count": len(clients),
                            "clients": clients
                        })
                        return
                
                self._json_response({"error": "Файл не найден в запросе"}, 400)
            else:
                # Прямая загрузка бинарных данных
                tmp_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "_upload.xlsx")
                with open(tmp_path, 'wb') as f:
                    f.write(body)
                
                clients = read_excel(tmp_path)
                _server_state["clients"] = clients
                os.remove(tmp_path)
                
                self._json_response({
                    "success": True,
                    "clients_count": len(clients),
                    "clients": clients
                })
        except Exception as e:
            self._json_response({"error": str(e)}, 500)
    
    def _handle_template(self, body):
        """Обновление шаблона SMS."""
        try:
            data = json.loads(body)
            _server_state["template"] = data.get("template", DEFAULT_TEMPLATE)
            self._json_response({"success": True})
        except Exception as e:
            self._json_response({"error": str(e)}, 400)
    
    def _handle_preview(self, body):
        """Предпросмотр сообщений для всех клиентов."""
        try:
            data = json.loads(body)
            template = data.get("template", _server_state["template"])
            previews = []
            for client in _server_state["clients"]:
                msg = render_message(template, client)
                previews.append({
                    "name": client["name"],
                    "phone": client["phone"],
                    "message": msg,
                    "chars": len(msg),
                    "sms_parts": (len(msg) // 70) + 1  # Кириллица = 70 символов на SMS
                })
            self._json_response({"previews": previews})
        except Exception as e:
            self._json_response({"error": str(e)}, 400)
    
    def _handle_provider(self, body):
        """Настройка SMS-провайдера."""
        try:
            data = json.loads(body)
            _server_state["provider_name"] = data.get("provider", "dryrun")
            _server_state["provider_config"] = data.get("config", {})
            self._json_response({"success": True, "provider": _server_state["provider_name"]})
        except Exception as e:
            self._json_response({"error": str(e)}, 400)
    
    def _handle_send(self, body):
        """Отправка SMS всем клиентам."""
        try:
            data = json.loads(body)
            template = data.get("template", _server_state["template"])
            provider_name = data.get("provider", _server_state["provider_name"])
            config = data.get("config", _server_state["provider_config"])
            
            provider = get_provider(provider_name, **config)
            
            # Проверяем баланс
            balance = provider.check_balance()
            
            results = send_notifications(
                _server_state["clients"],
                template,
                provider,
                delay=float(data.get("delay", 1.0))
            )
            
            _server_state["results"] = results
            
            ok = sum(1 for r in results if r['result']['success'])
            
            self._json_response({
                "success": True,
                "total": len(results),
                "sent": ok,
                "failed": len(results) - ok,
                "balance": balance,
                "results": [
                    {
                        "name": r["client"]["name"],
                        "phone": r["client"]["phone"],
                        "success": r["result"]["success"],
                        "error": r["result"]["error"],
                        "message_id": r["result"]["message_id"],
                    }
                    for r in results
                ]
            })
        except Exception as e:
            import traceback
            traceback.print_exc()
            self._json_response({"error": str(e)}, 500)
    
    def _json_response(self, data, status=200):
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Access-Control-Allow-Origin", "*")
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False, indent=2).encode('utf-8'))
    
    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()
    
    def log_message(self, format, *args):
        ts = datetime.now().strftime("%H:%M:%S")
        print(f"[{ts}] {args[0]}")


# ============================================================
# 🚀 CLI
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="🐣 SMS-уведомления о доставке (Азовский Инкубатор)")
    parser.add_argument("--file", "-f", help="Excel-файл с клиентами (.xlsx)")
    parser.add_argument("--template", "-t", help="Шаблон SMS (используйте {name}, {phone}, {order}, {address}, {time})")
    parser.add_argument("--provider", "-p", default="dryrun", choices=["smsru", "smsc", "dryrun"],
                        help="SMS-провайдер (по умолчанию: dryrun)")
    parser.add_argument("--api-key", help="API-ключ для SMS.ru")
    parser.add_argument("--login", help="Логин для SMSC.ru")
    parser.add_argument("--password", help="Пароль для SMSC.ru")
    parser.add_argument("--delay", type=float, default=1.0, help="Задержка между SMS (сек)")
    parser.add_argument("--dry-run", action="store_true", help="Тестовый режим (без реальной отправки)")
    parser.add_argument("--server", action="store_true", help="Запустить веб-панель")
    parser.add_argument("--port", type=int, default=8080, help="Порт веб-панели")
    
    args = parser.parse_args()
    
    # Режим сервера
    if args.server:
        print(f"🌐 Веб-панель: http://localhost:{args.port}")
        print(f"   Открой в браузере для управления рассылкой")
        server = HTTPServer(("0.0.0.0", args.port), NotifierHandler)
        try:
            server.serve_forever()
        except KeyboardInterrupt:
            print("\n⛔ Сервер остановлен")
        return
    
    # Режим CLI
    if not args.file:
        parser.error("Укажите файл: --file clients.xlsx")
    
    print("🐣 Система уведомлений — Азовский Инкубатор")
    print(f"   Файл: {args.file}")
    
    # Читаем клиентов
    clients = read_excel(args.file)
    print(f"   Найдено клиентов: {len(clients)}")
    
    if not clients:
        print("❌ Нет клиентов с валидными телефонами!")
        return
    
    # Шаблон
    template = args.template or DEFAULT_TEMPLATE
    
    # Провайдер
    if args.dry_run:
        provider = get_provider("dryrun")
        print("   Режим: 🧪 DRY RUN (SMS не отправляются)")
    elif args.provider == "smsru":
        if not args.api_key:
            parser.error("Для SMS.ru укажите --api-key")
        provider = get_provider("smsru", api_key=args.api_key)
        print("   Провайдер: SMS.ru")
    elif args.provider == "smsc":
        if not args.login or not args.password:
            parser.error("Для SMSC укажите --login и --password")
        provider = get_provider("smsc", login=args.login, password=args.password)
        print("   Провайдер: SMSC.ru")
    else:
        provider = get_provider("dryrun")
    
    # Баланс
    balance = provider.check_balance()
    if balance >= 0:
        print(f"   Баланс: {balance}₽")
    
    # Предпросмотр первого
    print(f"\n📝 Пример сообщения:")
    print(f"   {render_message(template, clients[0])}")
    print()
    
    # Подтверждение
    if not args.dry_run:
        confirm = input(f"Отправить {len(clients)} SMS? [y/N]: ").strip().lower()
        if confirm != 'y':
            print("⛔ Отменено.")
            return
    
    # Рассылка
    print("\n📤 Начинаю рассылку...\n")
    results = send_notifications(clients, template, provider, delay=args.delay)
    
    # Сохраняем лог
    log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")
    os.makedirs(log_dir, exist_ok=True)
    log_file = os.path.join(log_dir, f"send_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json")
    
    with open(log_file, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2, default=str)
    
    print(f"\n📋 Лог сохранён: {log_file}")


if __name__ == "__main__":
    main()
